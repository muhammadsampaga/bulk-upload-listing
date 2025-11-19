from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify, session
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import zipfile
import shutil
import json
import secrets
from openai import OpenAI
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'dev-secret-key')

def generate_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(16)
    return session['_csrf_token']

def validate_csrf_token(token):
    return token == session.get('_csrf_token')

app.jinja_env.globals['csrf_token'] = generate_csrf_token

# Template filter untuk format harga Indonesia (dengan titik)
def format_harga(value):
    """Format harga dengan pemisah titik (Indonesia format)"""
    if not value:
        return '-'
    try:
        harga = float(value)
        # Format dengan 0 desimal dan gunakan titik sebagai pemisah ribuan
        # Python format: {:,.0f} menggunakan comma, jadi kita ganti dengan titik
        formatted = "{:,.0f}".format(harga)
        return formatted.replace(',', '.')
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['format_harga'] = format_harga

UPLOAD_FOLDER = 'uploads'
EXCEL_FILE = os.path.join(UPLOAD_FOLDER, 'mass_upload_template.xlsx')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_zip_filename():
    """Generate ZIP filename with human-readable timestamp: upl_{DD-MMM-YYYY_HH-MM-SS}"""
    timestamp = datetime.now().strftime('%d-%b-%Y_%H-%M-%S')
    return os.path.join(UPLOAD_FOLDER, f'upl_{timestamp}.zip')

def get_current_zip_file():
    """Get the most recent ZIP file or generate new filename"""
    # Look for most recent upl_*.zip file
    try:
        files = [f for f in os.listdir(UPLOAD_FOLDER) if f.startswith('upl_') and f.endswith('.zip')]
        if files:
            files.sort(reverse=True)
            return os.path.join(UPLOAD_FOLDER, files[0])
    except:
        pass
    
    return get_zip_filename()

# Excel columns per property type with proper spacing in headers
EXCEL_COLUMNS_BY_TYPE = {
    'ruko': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'fasilitas ruko', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'rumah': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar tidur', 'kamar mandi', 'jumlah lantai', 'carport',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'apartemen': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar tidur', 'kamar mandi', 'nomor lantai',
        'sertifikat', 'kondisi perabotan', 'kondisi properti',
        'fasilitas lingkungan apartemen', 'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'tanah': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'satuan tanah', 'harga', 'satuan harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'sertifikat', 'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'villa': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar tidur', 'kamar mandi', 'jumlah lantai', 'carport',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'kost': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar tidur', 'kamar mandi', 'jumlah lantai', 'carport',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'hotel': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar tidur', 'kamar mandi', 'jumlah lantai', 'carport',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'pabrik': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai', 'carport',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'fasilitas pabrik', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'gudang': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'fasilitas gudang', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'perkantoran': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'fasilitas perkantoran', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'ruang usaha': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'fasilitas ruang usaha', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ],
    'gedung': [
        'no', 'tipe properti', 'kategori', 'jenis properti', 'id area',
        'luas tanah', 'luas bangunan', 'harga', 'periode sewa',
        'judul iklan', 'deskripsi iklan', 'gambar cover utama',
        'kamar mandi', 'jumlah lantai',
        'sertifikat', 'kondisi properti', 'kondisi perabotan',
        'selling point', 'Status',
        'aktivasi premier', 'aktivasi featured',
        'jadwal sundul', 'durasi sundul'
    ]
}

# Legacy global columns for backward compatibility (internal use only)
GLOBAL_COLUMNS = [
    'no', 'tipe_properti', 'kategori', 'jenis_properti', 'id_area',
    'harga', 'periode_sewa',
    'judul_iklan', 'deskripsi_iklan', 'gambar_cover_utama',
    'sertifikat', 'kondisi_properti', 'kondisi_perabotan',
    'selling_point', 'status', 'aktivasi_premier', 'aktivasi_featured',
    'jadwal_sundul', 'durasi_sundul'
]

# Specific columns per property type (internal use, with underscores)
SPECIFIC_COLUMNS = {
    'rumah': ['luas_tanah', 'luas_bangunan', 'kamar_tidur', 'kamar_mandi', 'jumlah_lantai', 'carport'],
    'apartemen': ['luas_bangunan', 'kamar_tidur', 'kamar_mandi', 'nomor_lantai', 'fasilitas_lingkungan'],
    'tanah': ['luas_tanah', 'satuan_tanah', 'satuan_harga'],
    'ruko': ['luas_tanah', 'luas_bangunan', 'kamar_mandi', 'jumlah_lantai', 'fasilitas_ruko'],
    'villa': ['luas_tanah', 'luas_bangunan', 'kamar_tidur', 'kamar_mandi', 'jumlah_lantai', 'carport'],
    'kost': ['luas_tanah', 'luas_bangunan', 'kamar_tidur', 'kamar_mandi', 'jumlah_lantai', 'carport'],
    'hotel': ['luas_tanah', 'luas_bangunan', 'kamar_tidur', 'kamar_mandi', 'jumlah_lantai', 'carport'],
    'pabrik': ['luas_tanah', 'luas_bangunan', 'kamar_mandi', 'jumlah_lantai', 'carport', 'fasilitas_pabrik'],
    'gudang': ['luas_tanah', 'luas_bangunan', 'kamar_mandi', 'jumlah_lantai', 'fasilitas_gudang'],
    'perkantoran': ['luas_tanah', 'luas_bangunan', 'kamar_mandi', 'jumlah_lantai', 'fasilitas_perkantoran'],
    'ruang_usaha': ['luas_tanah', 'luas_bangunan', 'kamar_mandi', 'jumlah_lantai', 'fasilitas_ruang_usaha']
}

def get_excel_columns(tipe_properti=None):
    """Get columns for Excel based on property type (with spaces in headers)"""
    if not tipe_properti:
        return EXCEL_COLUMNS_BY_TYPE.get('ruko', EXCEL_COLUMNS_BY_TYPE['ruko'])
    
    # Normalize the property type first
    tipe_normalized = normalize_tipe_properti(tipe_properti)
    
    # Check both formats
    if tipe_properti in EXCEL_COLUMNS_BY_TYPE:
        return EXCEL_COLUMNS_BY_TYPE[tipe_properti]
    if tipe_normalized in EXCEL_COLUMNS_BY_TYPE:
        return EXCEL_COLUMNS_BY_TYPE[tipe_normalized]
    
    # Default to ruko if not found
    return EXCEL_COLUMNS_BY_TYPE.get('ruko', EXCEL_COLUMNS_BY_TYPE['ruko'])

# Legacy column name for backward compatibility
EXCEL_COLUMNS = get_excel_columns()

ALLOWED_TIPE_PROPERTI = {
    'rumah', 'apartemen', 'ruko', 'tanah', 'villa', 'gedung',
    'kost', 'hotel', 'pabrik', 'gudang', 'perkantoran', 'ruang_usaha', 'ruang usaha'
}

# Load OpenAI API key: prefer environment variable, otherwise fall back to local config.py
CONFIG_OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY')
try:
    # optional local config file (create config.py with OPENAI_API_KEY = 'sk-...')
    from config import OPENAI_API_KEY as CONFIG_OPENAI_API_KEY  # type: ignore
except Exception:
    CONFIG_OPENAI_API_KEY = None

def get_openai_client():
    """Get or initialize OpenAI client (lazy loading for Render compatibility)"""
    api_key = os.environ.get('OPENAI_API_KEY') or CONFIG_OPENAI_API_KEY
    if not api_key:
        raise Exception("OpenAI API key tidak tersedia. Silakan tambahkan OPENAI_API_KEY di environment variables.")
    return OpenAI(api_key=api_key)

def generate_professional_listing(data, tipe_properti):
    """Generate professional judul_iklan and deskripsi_iklan using AI"""
    client = get_openai_client()
    
    try:
        # Build context from form data
        context_parts = [f"Tipe Properti: {tipe_properti}"]
        
        # Add key specifications
        specs = {
            'luas_tanah': data.get('luas_tanah', ''),
            'luas_bangunan': data.get('luas_bangunan', ''),
            'kamar_tidur': data.get('kamar_tidur', ''),
            'kamar_mandi': data.get('kamar_mandi', ''),
            'jumlah_lantai': data.get('jumlah_lantai', ''),
            'carport': data.get('carport', ''),
            'harga': data.get('harga', ''),
            'kategori': data.get('kategori', ''),
            'periode_sewa': data.get('periode_sewa', ''),
            'id_area': data.get('id_area', ''),
            'selling_point': data.get('selling_point', ''),
            'fasilitas_lingkungan': data.get('fasilitas_lingkungan', ''),
            'fasilitas_ruko': data.get('fasilitas_ruko', ''),
            'fasilitas_pabrik': data.get('fasilitas_pabrik', ''),
            'fasilitas_gudang': data.get('fasilitas_gudang', ''),
            'fasilitas_perkantoran': data.get('fasilitas_perkantoran', ''),
            'fasilitas_ruang_usaha': data.get('fasilitas_ruang_usaha', ''),
            'kondisi_properti': data.get('kondisi_properti', ''),
            'sertifikat': data.get('sertifikat', ''),
        }
        
        # Build context string
        for key, value in specs.items():
            if value:
                context_parts.append(f"{key}: {value}")
        
        context = "\n".join(context_parts)
        
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": """Kamu adalah agen properti profesional dan marketer terbaik yang sangat memahami cara menjual properti.
Tugasmu adalah membuat judul listing dan deskripsi yang menarik, profesional, dan persuasif.

ATURAN PENTING:
1. JUDUL LISTING (Maksimal 52 karakter):
   - Maksimalkan informasi penting di 52 karakter pertama
   - Format: [Tipe] [Key Info], [Harga/Keunggulan], [Lokasi Key]
   - Contoh: "Kost 14 Kamar, 400juta/Tahun, ada Laundry"

2. KALIMAT PEMBUKA (Maksimal 82 karakter):
   - Hook untuk membaca lebih lanjut
   - Nyatakan nilai utama properti
   - Contoh: "Kost di Denpasar, 3 lantai, 14 kamar, dengan Laundry, potensi investasi."

3. DESKRIPSI LENGKAP (Struktur: Keunggulan – Spesifikasi – Potensi – Legalitas – Call to Action):
   - GUNAKAN TEXT PLAIN SAJA, TIDAK PERLU MARKDOWN FORMATTING
   - JANGAN GUNAKAN: ###, **, -, 📍, 🌟, atau simbol/emoji lainnya
   - STRUKTUR: Gunakan line break dan paragraf untuk memisahkan bagian
   - Paragraf pembuka yang memikat
   - Spesifikasi Utama (list dengan tanda ketukan sederhana, contoh "- Luas Tanah: 100 m²")
   - Keunggulan khusus (format list sederhana)
   - Perhitungan finansial (jika relevan)
   - Legalitas
   - Call to Action yang kuat
   - Spacing dan line break untuk readability, bukan markdown

Respond ONLY dengan JSON object mengandung:
{
  "judul_iklan": "...",
  "kalimat_pembuka": "...",
  "deskripsi_iklan": "..."
}"""
                },
                {
                    "role": "user",
                    "content": f"Buatkan judul dan deskripsi profesional untuk properti dengan detail berikut:\n\n{context}"
                }
            ],
            response_format={"type": "json_object"}
        )
        
        result = json.loads(response.choices[0].message.content)
        return {
            'judul_iklan': result.get('judul_iklan', ''),
            'kalimat_pembuka': result.get('kalimat_pembuka', ''),
            'deskripsi_iklan': result.get('deskripsi_iklan', '')
        }
    except Exception as e:
        error_msg = f"Professional Listing Generator Error: {str(e)}"
        print(error_msg)
        return {'error': error_msg}

def parse_listing_with_ai(description):
    client = get_openai_client()
    
    try:
        # the newest OpenAI model is "gpt-5" which was released August 7, 2025.
        # do not change this unless explicitly requested by the user
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": """Kamu adalah asisten AI yang membantu mengekstrak informasi properti dari deskripsi teks.
Analisa deskripsi properti dan ekstrak informasi berikut dalam format JSON.
Jika informasi tidak tersedia atau tidak jelas, gunakan string kosong "".

FIELD GLOBAL (Semua Tipe Properti):
- tipe_properti: (rumah, apartemen, tanah, ruko, villa, kost, hotel, pabrik, gudang, perkantoran, ruang_usaha, gedung)
- kategori: (dijual, disewa)
- jenis_properti: (baru, second, aset_bank)
- id_area: nama lokasi/area
- harga: angka saja dalam rupiah (misal 5000000 untuk 5 juta)
- periode_sewa: (per hari, per bulan, per tahun) - hanya jika kategori=disewa - TANPA UNDERSCORE
- judul_iklan: max 52 karakter, menarik dan singkat, berisi info tipe + lokasi + harga
- deskripsi_iklan: PENTING! Buat deskripsi yang menarik dan profesional dengan struktur:
  * Kalimat pembuka singkat yang menarik (1-2 kalimat)
  * Spesifikasi detail (LT, LB, KT, KM, lantai, dll)
  * Keunggulan dan fasilitas utama
  * Lokasi strategis
  * Perihal transaksi jika ada
  Gunakan PLAIN TEXT SAJA - JANGAN gunakan markdown, emoji, atau format khusus. Gunakan newline untuk pemisah paragraf.
- sertifikat: (SHM, HGB, HP, HS, HGU, Adat, Girik, PPJB, Strata, Lainnya)
- kondisi_properti: (Bagus, Butuh Minim Renovasi, Butuh Renovasi Total, Terenovasi)
- kondisi_perabotan: (Furnished, Unfurnished, Semi Furnished)
- selling_point: keunggulan utama properti 2-3 kata
- status: (Tayang, Tunda)
- aktivasi_premier: (Yes, No)
- aktivasi_featured: (Yes, No)
- jadwal_sundul: (1 time daily, 2 times daily, 3 times daily)
- durasi_sundul: misal "7d" atau "30 hari"

FIELD KHUSUS PER TIPE PROPERTI:
Jika tipe_properti adalah RUMAH, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_tidur: angka
- kamar_mandi: angka
- jumlah_lantai: angka
- carport: angka

Jika tipe_properti adalah APARTEMEN, ekstrak juga:
- luas_bangunan: angka saja (m²)
- kamar_tidur: angka
- kamar_mandi: angka
- nomor_lantai: string
- fasilitas_lingkungan: string

Jika tipe_properti adalah TANAH, ekstrak juga:
- luas_tanah: angka saja (m²)
- satuan_tanah: (m2, are, ha)
- satuan_harga: (total, /m2, /are, /ha)

Jika tipe_properti adalah RUKO, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_mandi: angka
- jumlah_lantai: angka
- fasilitas_ruko: string

Jika tipe_properti adalah VILLA, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_tidur: angka
- kamar_mandi: angka
- jumlah_lantai: angka
- carport: angka

Jika tipe_properti adalah KOST, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_tidur: angka
- kamar_mandi: angka
- jumlah_lantai: angka
- carport: angka

Jika tipe_properti adalah HOTEL, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_tidur: angka
- kamar_mandi: angka
- jumlah_lantai: angka
- carport: angka

Jika tipe_properti adalah PABRIK, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_mandi: angka
- jumlah_lantai: angka
- carport: angka
- fasilitas_pabrik: string

Jika tipe_properti adalah GUDANG, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_mandi: angka
- jumlah_lantai: angka
- fasilitas_gudang: string

Jika tipe_properti adalah PERKANTORAN, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_mandi: angka
- jumlah_lantai: angka
- fasilitas_perkantoran: string

Jika tipe_properti adalah RUANG_USAHA, ekstrak juga:
- luas_tanah: angka saja (m²)
- luas_bangunan: angka saja (m²)
- kamar_mandi: angka
- jumlah_lantai: angka
- fasilitas_ruang_usaha: string

PENTING: 
- Pastikan SEMUA field yang relevan diisi dengan nilai yang tepat berdasarkan tipe properti!
- deskripsi_iklan HARUS berisi informasi lengkap dan menarik, bukan hanya ringkasan singkat
- Gunakan PLAIN TEXT tanpa markdown atau emoji
- Respond ONLY with valid JSON object, no additional text."""
                },
                {
                    "role": "user",
                    "content": f"Deskripsi properti: {description}"
                }
            ],
            response_format={"type": "json_object"}
        )
        
        result = json.loads(response.choices[0].message.content)
        return result
    except Exception as e:
        error_msg = f"AI Parser Error: {str(e)}"
        print(error_msg)
        return {'error': error_msg}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_tipe_properti(tipe_properti):
    """Normalize property type to internal format (with underscores)"""
    tipe_clean = tipe_properti.strip().lower()
    # Convert format with spaces to internal format with underscores
    mapping = {
        'ruang usaha': 'ruang_usaha'
    }
    return mapping.get(tipe_clean, tipe_clean)

def validate_tipe_properti(tipe_properti):
    tipe_clean = tipe_properti.strip().lower()
    if tipe_clean not in ALLOWED_TIPE_PROPERTI:
        return None
    # Return normalized version (with underscores for internal use)
    return normalize_tipe_properti(tipe_clean)

def init_excel(tipe_properti=None):
    """Initialize Excel file with appropriate columns for property type"""
    # Create main file if doesn't exist
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Properti"
        columns = get_excel_columns()
        for idx, col in enumerate(columns, 1):
            ws.cell(row=1, column=idx, value=col)
        wb.save(EXCEL_FILE)

def get_sheet_name(tipe_properti):
    """Get the sheet name based on property type (lowercase)"""
    # Normalize input first
    tipe_normalized = normalize_tipe_properti(tipe_properti)
    sheet_name_map = {
        'rumah': 'rumah',
        'apartemen': 'apartemen',
        'tanah': 'tanah',
        'ruko': 'ruko',
        'villa': 'villa',
        'kost': 'kost',
        'hotel': 'hotel',
        'pabrik': 'pabrik',
        'gudang': 'gudang',
        'perkantoran': 'perkantoran',
        'ruang_usaha': 'ruang_usaha',
        'gedung': 'gedung'
    }
    return sheet_name_map.get(tipe_normalized, 'data_properti')

def get_next_no(tipe_properti=None):
    """Get the next number for entry for specific property type"""
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE)
        
        if tipe_properti:
            # Get next number for specific property type
            tipe_normalized = normalize_tipe_properti(tipe_properti)
            sheet_name = get_sheet_name(tipe_normalized)
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_no = 1
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0]:
                        try:
                            max_no = max(max_no, int(row[0]) + 1)
                        except (ValueError, TypeError):
                            pass
                wb.close()
                return max_no
        
        # If no property type specified, default to 1
        wb.close()
        return 1
    except:
        return 1

def read_excel_data():
    """Read all data from all sheets in Excel"""
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE)
        data = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Only if 'no' is not empty
                    data.append(dict(zip(headers, row)))
        wb.close()  # Ensure workbook is closed
        return data
    except:
        return []

def create_header_mapping(tipe_properti):
    """Create mapping from Excel header (with spaces) to field name (with underscores)"""
    mapping = {
        'no': 'no',
        'tipe properti': 'tipe_properti',
        'kategori': 'kategori',
        'jenis properti': 'jenis_properti',
        'id area': 'id_area',
        'luas tanah': 'luas_tanah',
        'luas bangunan': 'luas_bangunan',
        'kamar tidur': 'kamar_tidur',
        'kamar mandi': 'kamar_mandi',
        'jumlah lantai': 'jumlah_lantai',
        'carport': 'carport',
        'nomor lantai': 'nomor_lantai',
        'fasilitas lingkungan': 'fasilitas_lingkungan',
        'fasilitas lingkungan apartemen': 'fasilitas_lingkungan',
        'satuan tanah': 'satuan_tanah',
        'satuan harga': 'satuan_harga',
        'harga': 'harga',
        'periode sewa': 'periode_sewa',
        'judul iklan': 'judul_iklan',
        'deskripsi iklan': 'deskripsi_iklan',
        'gambar cover utama': 'gambar_cover_utama',
        'sertifikat': 'sertifikat',
        'kondisi properti': 'kondisi_properti',
        'kondisi perabotan': 'kondisi_perabotan',
        'selling point': 'selling_point',
        'fasilitas ruko': 'fasilitas_ruko',
        'fasilitas pabrik': 'fasilitas_pabrik',
        'fasilitas gudang': 'fasilitas_gudang',
        'fasilitas perkantoran': 'fasilitas_perkantoran',
        'fasilitas ruang usaha': 'fasilitas_ruang_usaha',
        'Status': 'status',
        'status': 'status',
        'aktivasi premier': 'aktivasi_premier',
        'aktivasi featured': 'aktivasi_featured',
        'jadwal sundul': 'jadwal_sundul',
        'durasi sundul': 'durasi_sundul'
    }
    return mapping

def save_to_excel(data, tipe_properti=None):
    """Save data to Excel in the appropriate sheet based on property type"""
    init_excel()
    
    # Determine sheet name
    sheet_name = get_sheet_name(tipe_properti) if tipe_properti else "Data Properti"
    
    wb = load_workbook(EXCEL_FILE)
    
    # Create sheet if it doesn't exist
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        columns = get_excel_columns(tipe_properti)
        for idx, col in enumerate(columns, 1):
            ws.cell(row=1, column=idx, value=col)
    else:
        ws = wb[sheet_name]
    
    # Get columns from sheet header
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    # Create mapping for converting field names
    header_mapping = create_header_mapping(tipe_properti)
    
    # Build row data based on headers
    row_data = []
    for header in headers:
        # Try to get value using header name first (with spaces)
        value = data.get(header, '')
        # If not found, try mapping to underscore version and get from data
        if not value:
            field_name = header_mapping.get(header, header.lower().replace(' ', '_'))
            value = data.get(field_name, '')
        row_data.append(value)
    
    ws.append(row_data)
    wb.save(EXCEL_FILE)
    wb.close()  # Ensure file is properly closed

def update_excel_row(tipe_properti, no, updated_data):
    """Update an existing row in Excel"""
    init_excel()
    sheet_name = get_sheet_name(tipe_properti)
    
    wb = load_workbook(EXCEL_FILE)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    
    ws = wb[sheet_name]
    
    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    # Find and update row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value and str(row[0].value) == str(no):
            # Create mapping for converting field names
            header_mapping = create_header_mapping(tipe_properti)
            
            # Update each cell in the row
            for col_idx, header in enumerate(headers, 1):
                # Try to get value using header name first (with spaces)
                value = updated_data.get(header, '')
                # If not found, try mapping to underscore version
                if not value:
                    field_name = header_mapping.get(header, header.lower().replace(' ', '_'))
                    value = updated_data.get(field_name, '')
                
                ws.cell(row=row_idx, column=col_idx, value=value)
            break
    
    wb.save(EXCEL_FILE)
    wb.close()

def create_zip():
    """Create ZIP with structure: {tipe_properti}/{no}/ dengan Excel file"""
    zip_file = get_zip_filename()
    
    if os.path.exists(zip_file):
        os.remove(zip_file)
    
    with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Add Excel file - ensure it exists and is readable
        if os.path.exists(EXCEL_FILE):
            try:
                # Read Excel file into memory to avoid file lock issues
                with open(EXCEL_FILE, 'rb') as f:
                    excel_data = f.read()
                zipf.writestr('mass_upload_template.xlsx', excel_data)
            except Exception as e:
                print(f"Warning: Could not add Excel to ZIP: {str(e)}")
        
        # Walk through uploads folder and add image files with corrected path structure
        for root, dirs, files in os.walk(UPLOAD_FOLDER):
            for file in files:
                if file.endswith(('.png', '.jpg', '.jpeg', '.gif')):
                    file_path = os.path.join(root, file)
                    # Get relative path from UPLOAD_FOLDER
                    rel_path = os.path.relpath(file_path, UPLOAD_FOLDER)
                    # Don't include 'images' folder in the archive path
                    # Example: uploads/rumah/1/cover.jpg -> rumah/1/cover.jpg
                    parts = rel_path.split(os.sep)
                    # Filter out 'images' directory
                    filtered_parts = [p for p in parts if p != 'images']
                    arcname = os.sep.join(filtered_parts)
                    try:
                        zipf.write(file_path, arcname)
                    except Exception as e:
                        print(f"Warning: Could not add {file_path} to ZIP: {str(e)}")

@app.route('/')
def index():
    data = read_excel_data()
    return render_template('index.html', data=data)

@app.route('/parse-description', methods=['POST'])
def parse_description():
    try:
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'error': 'Request body harus berupa JSON yang valid'}), 400
        
        description = data.get('description', '')
        
        if not description:
            return jsonify({'error': 'Deskripsi tidak boleh kosong'}), 400
        
        parsed_data = parse_listing_with_ai(description)
        
        if 'error' in parsed_data:
            return jsonify(parsed_data), 500
        
        return jsonify(parsed_data)
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500

@app.route('/generate-listing', methods=['POST'])
def generate_listing():
    """Generate professional judul_iklan and deskripsi_iklan from form data"""
    try:
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'error': 'Request body harus berupa JSON yang valid'}), 400
        
        tipe_properti = data.get('tipe_properti', '')
        if not tipe_properti:
            return jsonify({'error': 'Tipe properti harus ditentukan'}), 400
        
        result = generate_professional_listing(data, tipe_properti)
        
        if 'error' in result:
            return jsonify(result), 500
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500

@app.route('/submit', methods=['POST'])
def submit():
    try:
        tipe_properti_raw = request.form.get('tipe_properti', '').strip()
        
        tipe_properti = validate_tipe_properti(tipe_properti_raw) if tipe_properti_raw else None
        if tipe_properti_raw and not validate_tipe_properti(tipe_properti_raw):
            flash(f'Tipe properti tidak valid: {tipe_properti_raw}', 'danger')
            return redirect(url_for('index'))
        
        if not tipe_properti:
            flash('Tipe properti harus dipilih', 'danger')
            return redirect(url_for('index'))
        
        # Get next number for this specific property type
        no = get_next_no(tipe_properti)
        
        property_folder = os.path.join(UPLOAD_FOLDER, tipe_properti, str(no))
        property_folder = os.path.abspath(property_folder)
        upload_folder_abs = os.path.abspath(UPLOAD_FOLDER)
        
        if not property_folder.startswith(upload_folder_abs):
            flash('Operasi tidak diizinkan: path tidak valid', 'danger')
            return redirect(url_for('index'))
        
        os.makedirs(property_folder, exist_ok=True)
        
        cover_filename = ''
        if 'gambar_cover' in request.files:
            cover_file = request.files['gambar_cover']
            if cover_file and cover_file.filename and allowed_file(cover_file.filename):
                ext = cover_file.filename.rsplit('.', 1)[1].lower()
                cover_filename = f'cover.{ext}'
                cover_file.save(os.path.join(property_folder, cover_filename))
        
        if 'gambar_lainnya' in request.files:
            other_files = request.files.getlist('gambar_lainnya')
            for idx, file in enumerate(other_files, 1):
                if file and file.filename and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f'foto{idx}.{ext}'
                    file.save(os.path.join(property_folder, filename))
        
        # Collect all form data including dynamic fields
        data = {
            'no': no,
            'tipe_properti': tipe_properti,
            'kategori': request.form.get('kategori', ''),
            'jenis_properti': request.form.get('jenis_properti', ''),
            'id_area': request.form.get('id_area', ''),
            'harga': request.form.get('harga', ''),
            'periode_sewa': request.form.get('periode_sewa', ''),
            'judul_iklan': request.form.get('judul_iklan', ''),
            'deskripsi_iklan': request.form.get('deskripsi_iklan', ''),
            'gambar_cover_utama': cover_filename,
            'sertifikat': request.form.get('sertifikat', ''),
            'kondisi_properti': request.form.get('kondisi_properti', ''),
            'kondisi_perabotan': request.form.get('kondisi_perabotan', ''),
            'selling_point': request.form.get('selling_point', ''),
            'status': request.form.get('status', ''),
            'aktivasi_premier': request.form.get('aktivasi_premier', ''),
            'aktivasi_featured': request.form.get('aktivasi_featured', ''),
            'jadwal_sundul': request.form.get('jadwal_sundul', ''),
            'durasi_sundul': request.form.get('durasi_sundul', '')
        }
        
        # Add specific fields for this property type
        if tipe_properti in SPECIFIC_COLUMNS:
            for field in SPECIFIC_COLUMNS[tipe_properti]:
                data[field] = request.form.get(field, '')
        
        save_to_excel(data, tipe_properti)
        create_zip()
        
        flash(f'Iklan tipe {tipe_properti.upper()} berhasil disimpan dan ZIP diperbarui', 'success')
    except Exception as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/download/excel')
def download_excel():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name='mass_upload_template.xlsx')
    flash('File Excel tidak ditemukan', 'warning')
    return redirect(url_for('index'))

@app.route('/download/zip')
def download_zip():
    zip_file = get_current_zip_file()
    if os.path.exists(zip_file):
        filename = os.path.basename(zip_file)
        return send_file(zip_file, as_attachment=True, download_name=filename)
    flash('File ZIP tidak ditemukan', 'warning')
    return redirect(url_for('index'))

@app.route('/load-data/<tipe>/<no>', methods=['GET'])
def load_data(tipe, no):
    """Load property data for editing"""
    try:
        tipe = validate_tipe_properti(tipe) if validate_tipe_properti(tipe) else tipe
        data = read_excel_data()
        
        # Find the data matching tipe and no
        for item in data:
            item_tipe = item.get('tipe properti') or item.get('tipe_properti', '')
            item_no = str(item.get('no', ''))
            
            if item_tipe.lower() == tipe.lower() and item_no == str(no):
                # Convert spaces to underscores in keys for JavaScript
                result = {}
                for key, value in item.items():
                    new_key = key.lower().replace(' ', '_')
                    result[new_key] = value
                return jsonify(result)
        
        return jsonify({'error': f'Data tidak ditemukan untuk {tipe} nomor {no}'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update-data', methods=['POST'])
def update_data():
    """Update existing property data"""
    try:
        tipe_properti_raw = request.form.get('tipe_properti', '').strip()
        no = request.form.get('no', '').strip()
        
        tipe_properti = validate_tipe_properti(tipe_properti_raw) if tipe_properti_raw else None
        if tipe_properti_raw and not validate_tipe_properti(tipe_properti_raw):
            flash(f'Tipe properti tidak valid: {tipe_properti_raw}', 'danger')
            return redirect(url_for('index'))
        
        if not tipe_properti or not no:
            flash('Tipe properti dan nomor harus ada', 'danger')
            return redirect(url_for('index'))
        
        # Read current data to delete old row
        data = read_excel_data()
        
        # Collect updated form data
        updated_data = {
            'no': no,
            'tipe_properti': tipe_properti,
            'kategori': request.form.get('kategori', ''),
            'jenis_properti': request.form.get('jenis_properti', ''),
            'id_area': request.form.get('id_area', ''),
            'harga': request.form.get('harga', ''),
            'periode_sewa': request.form.get('periode_sewa', ''),
            'judul_iklan': request.form.get('judul_iklan', ''),
            'deskripsi_iklan': request.form.get('deskripsi_iklan', ''),
            'sertifikat': request.form.get('sertifikat', ''),
            'kondisi_properti': request.form.get('kondisi_properti', ''),
            'kondisi_perabotan': request.form.get('kondisi_perabotan', ''),
            'selling_point': request.form.get('selling_point', ''),
            'status': request.form.get('status', ''),
            'aktivasi_premier': request.form.get('aktivasi_premier', ''),
            'aktivasi_featured': request.form.get('aktivasi_featured', ''),
            'jadwal_sundul': request.form.get('jadwal_sundul', ''),
            'durasi_sundul': request.form.get('durasi_sundul', '')
        }
        
        # Add specific fields for this property type
        if tipe_properti in SPECIFIC_COLUMNS:
            for field in SPECIFIC_COLUMNS[tipe_properti]:
                updated_data[field] = request.form.get(field, '')
        
        # Update Excel - remove old row and add new one
        update_excel_row(tipe_properti, no, updated_data)
        create_zip()
        
        flash(f'Data {tipe_properti} nomor {no} berhasil diperbarui', 'success')
    except Exception as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/reset-all', methods=['POST'])
def reset_all():
    csrf_token = request.form.get('csrf_token')
    if not csrf_token or not validate_csrf_token(csrf_token):
        flash('Invalid CSRF token. Operasi dibatalkan untuk keamanan.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Remove Excel file
        if os.path.exists(EXCEL_FILE):
            try:
                os.remove(EXCEL_FILE)
            except Exception as e:
                print(f"Warning: Could not remove Excel file: {str(e)}")
        
        # Only proceed if upload folder exists
        if os.path.exists(UPLOAD_FOLDER):
            # Remove property type folders
            try:
                for item in os.listdir(UPLOAD_FOLDER):
                    item_path = os.path.join(UPLOAD_FOLDER, item)
                    if os.path.isdir(item_path) and item in ALLOWED_TIPE_PROPERTI:
                        try:
                            shutil.rmtree(item_path)
                        except Exception as e:
                            print(f"Warning: Could not remove folder {item}: {str(e)}")
            except Exception as e:
                print(f"Warning: Error listing upload folder: {str(e)}")
            
            # Remove all ZIP files matching upl_*.zip pattern
            try:
                zip_files = [f for f in os.listdir(UPLOAD_FOLDER) if f.startswith('upl_') and f.endswith('.zip')]
                for zf in zip_files:
                    try:
                        os.remove(os.path.join(UPLOAD_FOLDER, zf))
                    except Exception as e:
                        print(f"Warning: Could not remove ZIP file {zf}: {str(e)}")
            except Exception as e:
                print(f"Warning: Error removing ZIP files: {str(e)}")
        
        # Ensure upload folder exists
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # Reinitialize Excel
        init_excel()
        
        flash('✅ Semua data berhasil direset! Nomor dan data kembali ke awal.', 'success')
    except Exception as e:
        flash(f'❌ Terjadi kesalahan saat reset: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
